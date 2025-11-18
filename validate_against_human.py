"""
Human Ground Truth Validation Script

This script loads human-coded scores from gold_standard.xlsx.
The Excel file has been pre-processed:
- First row contains column names (header=0)
- YT Link column is already forward-filled and unmerged
- Each video has 3 rows (one per evaluator: JPA, MMM, EA)

Dimensions evaluated:
1. Nuance (1-5): Oversimplification vs. Highly nuanced/contextualized
2. Order/Creativity (1-5): Order/control vs. Creativity/innovation
3. Prevention/Promotion (1-5): Prevention focus vs. Promotion focus
4. Contempt/Compassion (1-5): Contemptuous vs. Compassionate toward outgroups
5. Opinion/News (1-5): Opinion-based vs. Fact-based news reporting
"""

import pandas as pd
import numpy as np
import logging
import json
import os
import re
from scipy import stats
import matplotlib.pyplot as plt

try:
    import seaborn as sns
    HAS_SEABORN = True
except ImportError:
    HAS_SEABORN = False
    sns = None

from typing import Dict, List, Tuple, Optional
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# ============================================================================
# CONFIGURATION
# ============================================================================

GOLD_STANDARD_PATH = "gold_standard.xlsx"
OUTPUT_DIR = "validation_results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Dimension mappings - all possible column name patterns for each dimension
DIMENSION_PATTERNS = {
    'nuance': [
        'nuance', 'nuance score', 'text - nuance score', 'video - nuance score',
        'nuance/ oversimplification', 'n_rationale'
    ],
    'order_creativity': [
        'order', 'creativity', 'order/creativity', 'order/ creativity',
        'text - order/ creativity score', 'text - order/creativity score',
        'creativity_vs_order', 'c-o_rationale', 'o/c_rationale', 'creativity rational'
    ],
    'prevention_promotion': [
        'prevention', 'promotion', 'prevention/promotion', 'prevention/ promotion',
        'text - prevention/ promotion score', 'text - prevention/promotion score',
        'safety_vs_threat', 'p_rationale', 'prevention/promotion score'
    ],
    'contempt_compassion': [
        'contempt', 'compassion', 'contempt/compassion', 'contempt/ compassion',
        'text - contempt/ compassion score', 'text - contempt/compassion score',
        'respect_vs_contempt', 'r-c_rationale', 'c/c_rationale', 'c/c_rationale',
        'text - contempt/compassion score'
    ],
    'opinion_news': [
        'opinion', 'news', 'reporting', 'opinion vs. news', 'opinion/ news',
        'text - opinion/ news score', 'text - opinion vs. news score',
        'reporting_vs_opinion', 'r-o_rationale', 'o/n_rationale', 'o/n_rationale'
    ]
}

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def extract_video_id_from_link(link) -> Optional[str]:
    """
    Extract YouTube video ID from various formats OR create unique identifier from title.
    
    Handles:
    - YouTube URLs (extracts 11-char video ID)
    - Video titles (creates normalized identifier)
    - Hyperlinks (extracts URL first, then processes)
    - Mixed text (tries to extract ID, falls back to normalized title)
    
    The same video link/title will always produce the same identifier,
    which is important since each video appears 3 times (once per evaluator).
    """
    if pd.isna(link):
        return None
    
    link_str = str(link).strip()
    
    if not link_str or len(link_str) < 3:
        return None
    
    # First, try to extract video ID from URL patterns
    # Use utils function if available
    try:
        from utils import get_video_id
        video_id = get_video_id(link_str)
        if video_id and len(video_id) == 11:
            return video_id
    except:
        pass
    
    # Try URL patterns directly
    url_patterns = [
        r'youtube\.com/watch\?v=([a-zA-Z0-9_-]{11})',
        r'youtu\.be/([a-zA-Z0-9_-]{11})',
        r'youtube\.com/embed/([a-zA-Z0-9_-]{11})',
        r'youtube\.com/v/([a-zA-Z0-9_-]{11})',
    ]
    
    for pattern in url_patterns:
        match = re.search(pattern, link_str, re.IGNORECASE)
        if match:
            vid_id = match.group(1)
            if len(vid_id) == 11:  # YouTube IDs are 11 characters
                return vid_id
    
    # Check if it's just an 11-character ID by itself (standalone video ID)
    standalone_id_match = re.match(r'^([a-zA-Z0-9_-]{11})$', link_str)
    if standalone_id_match:
        return standalone_id_match.group(1)
    
    # If no URL/ID found, create a unique identifier from the title/link text
    # This handles cases where YT Link contains video titles or other text
    # Normalize: remove special chars, lowercase, take meaningful portion
    normalized = re.sub(r'[^\w\s]', '', link_str.lower())  # Remove special chars
    normalized = re.sub(r'\s+', '_', normalized.strip())   # Replace spaces with underscores
    
    # Take first 50 characters to create identifier (long enough to be unique)
    # But ensure it's at least 10 chars if possible
    if len(normalized) >= 10:
        identifier = normalized[:50]
    elif len(normalized) >= 5:
        identifier = normalized
    else:
        # Too short, might not be unique - return None to skip
        return None
    
    return identifier

def clean_score_value(value) -> Optional[float]:
    """
    Convert various score formats to float (1-5 scale).
    Handles: single numbers (3), ranges (3-4), text with numbers, dates, etc.
    For ranges, returns the midpoint.
    """
    if pd.isna(value):
        return None
    
    # Skip if it's a date/datetime
    if isinstance(value, (pd.Timestamp, datetime)):
        return None
    
    value_str = str(value).strip()
    
    # Handle ranges like "3-4" or "2-3" - take midpoint
    range_match = re.search(r'([0-9]+\.?[0-9]*)\s*[-–—]\s*([0-9]+\.?[0-9]*)', value_str)
    if range_match:
        low = float(range_match.group(1))
        high = float(range_match.group(2))
        score = (low + high) / 2.0
        if 0 <= score <= 5:
            return score
        return None
    
    # Try to extract single number
    if isinstance(value, (int, float)):
        score = float(value)
    else:
        # Extract first number found
        match = re.search(r'([0-9]+\.?[0-9]*)', value_str)
        if match:
            score = float(match.group(1))
        else:
            return None
    
    # Validate range (1-5, but allow 0-5 for flexibility)
    if 0 <= score <= 5:
        return score
    return None

def find_score_columns(df: pd.DataFrame, dimension: str) -> List[str]:
    """Find all columns that match a dimension's patterns."""
    score_cols = []
    patterns = DIMENSION_PATTERNS.get(dimension, [])
    
    for col in df.columns:
        col_str = str(col).lower().strip()
        # Skip rationale columns
        if 'rationale' in col_str or 'rational' in col_str:
            continue
        # Check if column matches any pattern and contains 'score'
        if any(pattern in col_str for pattern in patterns) and 'score' in col_str:
            score_cols.append(col)
    
    return score_cols

# ============================================================================
# LOAD AND EXTRACT DATA
# ============================================================================

def load_sheet_data(sheet_name: str) -> pd.DataFrame:
    """Load a sheet with header=0 (first row is column names)."""
    logging.info(f"Loading sheet: {sheet_name}")
    df = pd.read_excel(GOLD_STANDARD_PATH, sheet_name=sheet_name, header=0)
    
    # Remove completely empty rows
    df = df.dropna(how='all').reset_index(drop=True)
    
    logging.info(f"  Loaded {len(df)} rows, {len(df.columns)} columns")
    return df

def extract_video_id_from_excel_hyperlink(sheet_name: str, row_idx: int, col_name: str) -> Optional[str]:
    """
    Extract YouTube video ID from Excel cell hyperlink (not just text).
    Uses openpyxl to access actual hyperlink target.
    """
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(GOLD_STANDARD_PATH, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        
        ws = wb[sheet_name]
        
        # Find column index
        header_row = 1  # Row 1 is headers (0-indexed would be row 2, but openpyxl is 1-indexed)
        col_idx = None
        for cell in ws[header_row]:
            if str(cell.value).lower().strip() == col_name.lower().strip():
                col_idx = cell.column
                break
        
        if not col_idx:
            return None
        
        # Get cell (row_idx + 2 because: row 0 is info, row 1 is header, data starts at row 2)
        data_row = row_idx + 2
        cell = ws.cell(row=data_row, column=col_idx)
        
        # Check if cell has hyperlink
        if cell.hyperlink and cell.hyperlink.target:
            hyperlink_url = cell.hyperlink.target
            # Extract video ID from hyperlink URL
            video_id = extract_video_id_from_link(hyperlink_url)
            if video_id and len(video_id) == 11:  # Valid YouTube ID
                return video_id
        
        # Fallback to cell value
        if cell.value:
            return extract_video_id_from_link(str(cell.value))
        
        return None
    except Exception as e:
        logging.debug(f"  Could not extract hyperlink from Excel: {e}")
        return None

def extract_all_scores_from_sheet(sheet_name: str) -> pd.DataFrame:
    """Extract all scores from a sheet for all dimensions."""
    df = load_sheet_data(sheet_name)
    
    if len(df) == 0:
        return pd.DataFrame()
    
    # Find YT Link and Evaluator columns
    yt_link_col = None
    evaluator_col = None
    
    for col in df.columns:
        col_str = str(col).lower().strip()
        if 'link' in col_str and ('yt' in col_str or 'youtube' in col_str):
            yt_link_col = col
        elif 'evaluator' in col_str or 'initial' in col_str:
            evaluator_col = col
    
    if not yt_link_col or not evaluator_col:
        logging.warning(f"  Could not find YT Link or Evaluator column in {sheet_name}")
        return pd.DataFrame()
    
    logging.info(f"  Found columns - YT Link: {yt_link_col}, Evaluator: {evaluator_col}")
    
    # Extract all scores for all dimensions
    results = []
    skipped_rows = 0
    
    for idx, row in df.iterrows():
        # Try to get video ID from Excel hyperlink first
        video_id = extract_video_id_from_excel_hyperlink(sheet_name, int(idx), yt_link_col)
        
        # Fallback to text extraction if hyperlink method failed
        if not video_id or len(video_id) != 11:
            link_value = row.get(yt_link_col)
            video_id = extract_video_id_from_link(link_value)
        
        if not video_id:
            skipped_rows += 1
            # Only log first few skipped rows to avoid spam
            if skipped_rows <= 3:
                link_value = row.get(yt_link_col)
                logging.debug(f"  Row {idx}: Could not extract video ID from: {str(link_value)[:60]}")
            continue
        
        # Store both normalized ID (for matching) and actual YouTube ID (for transcript fetching)
        normalized_id = video_id  # Keep for backward compatibility
        youtube_id = video_id if video_id and len(video_id) == 11 else None
        
        # Get evaluator
        evaluator = str(row.get(evaluator_col)).strip() if pd.notna(row.get(evaluator_col)) else None
        if not evaluator or evaluator.lower() in ['nan', 'none', '']:
            skipped_rows += 1
            continue
        evaluator = evaluator.upper()
        
        # Extract scores for each dimension
        row_scores = {
            'video_id': normalized_id,  # Normalized ID for matching
            'youtube_id': youtube_id,   # Actual YouTube ID for transcript fetching
            'evaluator': evaluator,
            'sheet': sheet_name
        }
        
        # Check each dimension
        for dimension in DIMENSION_PATTERNS.keys():
            score_cols = find_score_columns(df, dimension)
            
            scores = []
            for score_col in score_cols:
                if score_col in df.columns:
                    score_val = clean_score_value(row.get(score_col))
                    if score_val is not None:
                        scores.append(score_val)
            
            # Use average if multiple scores, or single score
            if scores:
                final_score = float(np.mean(scores))
                if not np.isnan(final_score):
                    row_scores[f'{dimension}_score'] = final_score
        
        # Only add if we have at least one score
        if any(k.endswith('_score') for k in row_scores.keys()):
            results.append(row_scores)
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        unique_videos = result_df['video_id'].nunique() if 'video_id' in result_df.columns else 0
        logging.info(f"  Extracted {len(result_df)} rows with scores from {unique_videos} unique videos")
        if skipped_rows > 0:
            logging.info(f"  Skipped {skipped_rows} rows (no valid video ID or evaluator)")
    else:
        logging.warning(f"  No scores extracted from {sheet_name}")
    
    return result_df

def load_all_human_scores() -> Tuple[pd.DataFrame, Dict]:
    """
    Load all human scores from all sheets.
    Returns aggregated DataFrame (gold standard = average across evaluators) and notes.
    """
    logging.info("="*80)
    logging.info("LOADING ALL HUMAN SCORES FROM GOLD STANDARD")
    logging.info("="*80)
    
    xl = pd.ExcelFile(GOLD_STANDARD_PATH)
    notes = {
        'missing_data': [],
        'inconsistencies': [],
        'evaluators': set(),
        'dimensions_found': set(),
        'sheets_processed': []
    }
    
    all_scores = []
    
    # Process each sheet
    for sheet_name in xl.sheet_names:
        if sheet_name == 'Form Responses 1':
            continue  # Skip form responses sheet
        
        try:
            df_sheet = extract_all_scores_from_sheet(sheet_name)
            if len(df_sheet) > 0:
                all_scores.append(df_sheet)
                notes['sheets_processed'].append(sheet_name)
                
                # Track evaluators and dimensions
                if 'evaluator' in df_sheet.columns:
                    notes['evaluators'].update(df_sheet['evaluator'].dropna().unique())
                
                score_cols = [c for c in df_sheet.columns if c.endswith('_score')]
                notes['dimensions_found'].update([c.replace('_score', '') for c in score_cols])
                
                logging.info(f"  ✓ {sheet_name}: {len(df_sheet)} rows")
        except Exception as e:
            notes['inconsistencies'].append(f"Error processing {sheet_name}: {str(e)}")
            logging.warning(f"  ✗ Error processing {sheet_name}: {e}")
    
    if not all_scores:
        raise ValueError("No human scores could be extracted from gold_standard.xlsx")
    
    # Combine all dataframes
    human_df = pd.concat(all_scores, ignore_index=True)
    
    # Keep individual scores for inter-rater analysis
    human_df_individual = human_df.copy()
    
    # Aggregate by video_id (average across evaluators = gold standard)
    score_cols = [col for col in human_df.columns if col.endswith('_score')]
    
    if len(score_cols) == 0:
        raise ValueError("No score columns found after combining sheets")
    
    # Build aggregation dictionary
    agg_dict = {col: 'mean' for col in score_cols}
    
    # Keep YouTube ID (take first non-null value, prefer 11-char IDs)
    if 'youtube_id' in human_df.columns:
        def first_valid_youtube_id(x):
            valid = x.dropna()
            if len(valid) > 0:
                # Prefer 11-char YouTube IDs
                for val in valid:
                    if pd.notna(val) and len(str(val)) == 11:
                        return str(val)
                return str(valid.iloc[0])
            return None
        agg_dict['youtube_id'] = first_valid_youtube_id
    
    # Keep evaluator info
    if 'evaluator' in human_df.columns:
        def join_unique(x):
            return ','.join(sorted(x.dropna().astype(str).unique()))
        agg_dict['evaluator'] = join_unique
    
    # Keep sheet info
    if 'sheet' in human_df.columns:
        def join_unique_sheet(x):
            return ','.join(x.dropna().astype(str).unique())
        agg_dict['sheet'] = join_unique_sheet
    
    # Create gold standard (averaged across evaluators)
    human_df_agg = human_df.groupby('video_id').agg(agg_dict).reset_index()
    
    # Track info
    notes['evaluators'] = sorted(list(notes['evaluators']))
    notes['dimensions_found'] = sorted(list(notes['dimensions_found']))
    notes['individual_scores_df'] = human_df_individual
    
    logging.info(f"\n✓ Loaded {len(human_df_agg)} unique videos")
    logging.info(f"✓ Dimensions found: {notes['dimensions_found']}")
    logging.info(f"✓ Evaluators: {notes['evaluators']}")
    logging.info(f"✓ Total individual ratings: {len(human_df_individual)}")
    
    return human_df_agg, notes

# ============================================================================
# DATA EXPLORATION AND SUMMARY
# ============================================================================

def create_human_metrics_summary(human_df: pd.DataFrame, notes: Dict, output_dir: str):
    """Create comprehensive summary statistics and visualizations."""
    logging.info("\n" + "="*80)
    logging.info("CREATING HUMAN METRICS SUMMARY")
    logging.info("="*80)
    
    score_cols = [col for col in human_df.columns if col.endswith('_score')]
    
    if len(score_cols) == 0:
        logging.warning("No score columns found for summary")
        return
    
    # 1. Basic statistics for each dimension
    summary_stats = []
    for col in score_cols:
        dimension = col.replace('_score', '')
        scores = human_df[col].dropna()
        
        if len(scores) > 0:
            summary_stats.append({
                'Dimension': dimension,
                'N': len(scores),
                'Mean': round(scores.mean(), 3),
                'Std': round(scores.std(), 3),
                'Min': round(scores.min(), 3),
                'Max': round(scores.max(), 3),
                'Median': round(scores.median(), 3),
                'Q1': round(scores.quantile(0.25), 3),
                'Q3': round(scores.quantile(0.75), 3),
                'Missing': len(human_df) - len(scores)
            })
    
    summary_df = pd.DataFrame(summary_stats)
    summary_path = os.path.join(output_dir, 'human_metrics_summary.csv')
    summary_df.to_csv(summary_path, index=False)
    logging.info(f"\n✓ Summary statistics saved to: {summary_path}")
    print("\n" + "="*80)
    print("HUMAN METRICS SUMMARY")
    print("="*80)
    print(summary_df.to_string(index=False))
    
    # 2. Pivot table: scores by dimension
    if len(score_cols) > 1:
        pivot_data = human_df[['video_id'] + score_cols].set_index('video_id')
        pivot_path = os.path.join(output_dir, 'human_scores_pivot.csv')
        pivot_data.to_csv(pivot_path)
        logging.info(f"✓ Pivot table saved to: {pivot_path}")
    
    # 3. Correlation matrix between dimensions
    if len(score_cols) > 1:
        corr_matrix = human_df[score_cols].corr()
        corr_path = os.path.join(output_dir, 'human_dimensions_correlation.csv')
        corr_matrix.to_csv(corr_path)
        logging.info(f"✓ Correlation matrix saved to: {corr_path}")
        
        # Visualize correlation
        if HAS_SEABORN:
            plt.figure(figsize=(10, 8))
            sns.heatmap(corr_matrix, annot=True, fmt='.3f', cmap='coolwarm', center=0,
                       square=True, linewidths=1, cbar_kws={"shrink": 0.8})
            plt.title('Correlation Between Human Rating Dimensions', fontsize=14, fontweight='bold')
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, 'human_dimensions_correlation.png'), 
                       dpi=300, bbox_inches='tight')
            plt.close()
            logging.info(f"✓ Correlation heatmap saved")
    
    # 4. Distribution plots for each dimension
    n_dims = len(score_cols)
    if n_dims > 0:
        n_rows = (n_dims + 1) // 2
        fig, axes = plt.subplots(n_rows, 2, figsize=(14, 4 * n_rows))
        if n_dims == 1:
            axes = [axes]
        else:
            axes = axes.flatten()
        
        for idx, col in enumerate(score_cols):
            if idx >= len(axes):
                break
            ax = axes[idx]
            dimension = col.replace('_score', '').replace('_', ' ').title()
            scores = human_df[col].dropna()
            
            ax.hist(scores, bins=20, edgecolor='black', alpha=0.7)
            ax.axvline(scores.mean(), color='red', linestyle='--', linewidth=2, 
                      label=f'Mean = {scores.mean():.2f}')
            ax.set_xlabel('Score (1-5)', fontsize=11)
            ax.set_ylabel('Frequency', fontsize=11)
            ax.set_title(f'{dimension} Distribution', fontsize=12, fontweight='bold')
            ax.legend()
            ax.grid(True, alpha=0.3)
        
        # Hide extra subplots
        for idx in range(len(score_cols), len(axes)):
            axes[idx].set_visible(False)
        
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, 'human_scores_distributions.png'), 
                   dpi=300, bbox_inches='tight')
        plt.close()
        logging.info(f"✓ Distribution plots saved")
    
    # 5. Missing data report
    missing_report = []
    for col in score_cols:
        missing_count = human_df[col].isna().sum()
        missing_pct = (missing_count / len(human_df)) * 100
        missing_report.append({
            'Dimension': col.replace('_score', ''),
            'Missing_Count': missing_count,
            'Missing_Percent': round(missing_pct, 2),
            'Available_Count': len(human_df) - missing_count
        })
    
    missing_df = pd.DataFrame(missing_report)
    missing_path = os.path.join(output_dir, 'missing_data_report.csv')
    missing_df.to_csv(missing_path, index=False)
    logging.info(f"\n✓ Missing data report saved to: {missing_path}")
    print("\n" + "="*80)
    print("MISSING DATA REPORT")
    print("="*80)
    print(missing_df.to_string(index=False))
    
    # 6. Save notes
    notes_path = os.path.join(output_dir, 'data_quality_notes.json')
    with open(notes_path, 'w') as f:
        json.dump({
            'missing_data': notes.get('missing_data', []),
            'inconsistencies': notes.get('inconsistencies', []),
            'evaluators': notes.get('evaluators', []),
            'dimensions_found': notes.get('dimensions_found', []),
            'sheets_processed': notes.get('sheets_processed', []),
            'total_videos': len(human_df),
            'timestamp': datetime.now().isoformat(),
            'notes': [
                "Gold standard created by averaging scores across 3 evaluators (JPA, MMM, EA).",
                "Each video was evaluated by multiple human raters.",
                "Some dimensions may be missing for certain videos.",
                "Scores are on a 1-5 scale for all dimensions."
            ]
        }, f, indent=2)
    logging.info(f"✓ Data quality notes saved to: {notes_path}")

# ============================================================================
# INTER-RATER RELIABILITY ANALYSIS
# ============================================================================

def analyze_inter_rater_reliability(human_df_individual: pd.DataFrame, output_dir: str):
    """Analyze inter-rater reliability across evaluators."""
    logging.info("\n" + "="*80)
    logging.info("INTER-RATER RELIABILITY ANALYSIS")
    logging.info("="*80)
    
    if 'evaluator' not in human_df_individual.columns:
        logging.warning("No evaluator column found. Skipping inter-rater analysis.")
        return None
    
    score_cols = [col for col in human_df_individual.columns if col.endswith('_score')]
    
    if len(score_cols) == 0:
        logging.warning("No score columns found for inter-rater analysis.")
        return None
    
    reliability_results = []
    
    for score_col in score_cols:
        dimension = score_col.replace('_score', '')
        
        # Pivot to get evaluators as columns
        pivot_df = human_df_individual.pivot_table(
            index='video_id',
            columns='evaluator',
            values=score_col,
            aggfunc='first'
        )
        
        evaluators = [col for col in pivot_df.columns if pd.notna(col)]
        
        if len(evaluators) < 2:
            logging.warning(f"  {dimension}: Need at least 2 evaluators")
            continue
        
        # Calculate correlations
        corr_matrix = pivot_df[evaluators].corr()
        
        # Average pairwise correlation
        pairwise_corrs = []
        for i, eval1 in enumerate(evaluators):
            for eval2 in evaluators[i+1:]:
                if eval1 in corr_matrix.index and eval2 in corr_matrix.columns:
                    corr_val = corr_matrix.loc[eval1, eval2]
                    if pd.notna(corr_val):
                        pairwise_corrs.append(corr_val)
        
        avg_correlation = np.mean(pairwise_corrs) if pairwise_corrs else None
        
        # Agreement within 1 point
        agreement_within_1 = []
        for eval1 in evaluators:
            for eval2 in evaluators:
                if eval1 != eval2:
                    diff = (pivot_df[eval1] - pivot_df[eval2]).abs()
                    within_1 = (diff <= 1.0).sum() / len(diff.dropna()) if len(diff.dropna()) > 0 else 0
                    agreement_within_1.append(within_1)
        
        avg_agreement = np.mean(agreement_within_1) if agreement_within_1 else None
        
        reliability_results.append({
            'dimension': dimension,
            'n_videos': len(pivot_df),
            'evaluators': ','.join(evaluators),
            'avg_pairwise_correlation': round(float(avg_correlation), 4) if avg_correlation is not None and not np.isnan(avg_correlation) else None,
            'avg_agreement_within_1pt': round(float(avg_agreement * 100), 2) if avg_agreement is not None and not np.isnan(avg_agreement) else None
        })
        
        if avg_correlation is not None and avg_agreement is not None:
            logging.info(f"  {dimension}: Avg correlation = {avg_correlation:.4f}, Agreement = {avg_agreement*100:.2f}%")
    
    if reliability_results:
        reliability_df = pd.DataFrame(reliability_results)
        reliability_path = os.path.join(output_dir, 'inter_rater_reliability.csv')
        reliability_df.to_csv(reliability_path, index=False)
        logging.info(f"\n✓ Inter-rater reliability saved to: {reliability_path}")
        
        print("\n" + "="*80)
        print("INTER-RATER RELIABILITY")
        print("="*80)
        print(reliability_df.to_string(index=False))
        
        return reliability_df
    
    return None

# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Extract and validate human scores from gold_standard.xlsx. "
                    "Creates gold standard by averaging scores across evaluators."
    )
    parser.add_argument("--gold-standard", type=str, default=GOLD_STANDARD_PATH,
                       help="Path to gold_standard.xlsx file")
    parser.add_argument("--skip-exploration", action="store_true",
                       help="Skip data exploration and summary statistics")
    
    args = parser.parse_args()
    
    GOLD_STANDARD_PATH = args.gold_standard
    
    try:
        # Load human scores
        human_df, notes = load_all_human_scores()
        
        # Save cleaned human scores (gold standard)
        human_scores_path = os.path.join(OUTPUT_DIR, 'human_scores_cleaned.csv')
        human_df.to_csv(human_scores_path, index=False)
        logging.info(f"\n✓ Gold standard saved to: {human_scores_path}")
        
        # Data exploration and summary
        if not args.skip_exploration:
            create_human_metrics_summary(human_df, notes, OUTPUT_DIR)
            
            # Inter-rater reliability analysis
            if 'individual_scores_df' in notes:
                analyze_inter_rater_reliability(notes['individual_scores_df'], OUTPUT_DIR)
        
        logging.info(f"\n✅ Human scores extraction and validation complete!")
        logging.info(f"   Gold standard (averaged across evaluators) saved to: {human_scores_path}")
        logging.info(f"   All results saved to {OUTPUT_DIR}/")
        logging.info(f"\n   Next step: Run your 4 models on these video transcripts,")
        logging.info(f"   then compare model scores against the gold standard averages.")
        
    except Exception as e:
        logging.error(f"\n❌ Validation failed: {e}")
        import traceback
        traceback.print_exc()
        raise
